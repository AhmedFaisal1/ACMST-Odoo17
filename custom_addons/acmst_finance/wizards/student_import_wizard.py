# -*- coding: utf-8 -*-
from odoo import models, fields, _, api
from odoo.exceptions import UserError

import base64
from io import BytesIO
import openpyxl
import logging
import time

_logger = logging.getLogger(__name__)

BATCH_CREATE = 1000
COMMIT_EVERY = 1000
LOG_EVERY_ROWS = 1000


class StudentImportWizard(models.TransientModel):
    _name = "acmst.student.import.wizard"
    _description = "Import Students from Excel"

    file = fields.Binary(string="File", required=True)
    filename = fields.Char(string="Filename")

    # ---------- helpers ----------
    @staticmethod
    def _to_text(val):
        if val is None:
            return False
        if isinstance(val, float):
            return str(int(val)) if val.is_integer() else str(val)
        return str(val).strip()

    @staticmethod
    def _map_sex(val):
        """
        Map Excel cell to selection value.
        In your sheet: 2 => male (m), 1 => female (f).
        Return 'm'/'f' (or False if empty/unknown).
        """
        if val is None:
            return False

        # رقم (int/float)
        if isinstance(val, (int, float)):
            try:
                n = int(val)
                if n == 2:
                    return "m"
                if n == 1:
                    return "f"
            except Exception:
                pass

        # كنص
        s = str(val).strip()
        if not s:
            return False

        sl = s.lower()

        male_codes = {"2", "٢", "2.0", "m", "male", "ذكر", "زكر"}
        female_codes = {"1", "١", "1.0", "f", "female", "انثى", "أنثى", "انثي"}

        if sl in male_codes:
            return "m"
        if sl in female_codes:
            return "f"

        # محاولة أخيرة لو نص رقم
        try:
            n = int(float(s))
            if n == 2:
                return "m"
            if n == 1:
                return "f"
        except Exception:
            pass

        return False

    # ---------- main ----------
    def action_import(self):
        self.ensure_one()
        t0 = time.time()
        _logger.info("ACMST import: START (filename=%s, b64_len=%s)", self.filename, len(self.file or b""))

        if not self.file:
            raise UserError(_("Please choose an .xlsx file."))
        fname = (self.filename or "").lower()
        if not fname.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            raise UserError(_("Only Excel OOXML files are supported: .xlsx, .xlsm, .xltx, .xltm"))

        # Decode + open workbook (read_only = fast/low memory)
        try:
            raw = base64.b64decode(self.file)
        except Exception:
            raise UserError(_("Could not decode the uploaded file. Please re-upload."))
        try:
            wb = openpyxl.load_workbook(filename=BytesIO(raw), data_only=True, read_only=True)
        except Exception as e:
            raise UserError(_("Invalid or corrupted Excel file '%s': %s") % (self.filename, e))
        ws = wb.active
        _logger.info("ACMST import: workbook opened, rows=%s, cols=%s", ws.max_row, ws.max_column)

        # Header
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(v).strip().upper() if v is not None else "" for v in header_row]
        colmap = {
            "FRMNO": "frmno",
            "FAC": "fac",
            "UNIV_ID": "univ_id",
            "N1": "n1",
            "N2": "n2",
            "N3": "n3",
            "N4": "n4",
            "SCNAME": "scname",
            "GOBNO": "gobno",
            "FACNAME": "facname",
            "GOBOLS": "gobols",
            "YEAR": "year",
            "NATIONAL_ID": "national_id",
            "SEX": "sex",
            "UNIVERSITY": "university",
        }
        missing = [h for h in colmap if h not in header]
        if missing:
            raise UserError(_("Missing columns in header: %s") % ", ".join(missing))
        idx = {h: header.index(h) for h in colmap}
        _logger.info("ACMST import: headers OK (%.2fs)", time.time() - t0)

        # --- Stream rows, de-duplicate by FRMNO in-file ---
        unique_by_frmno = {}   # frmno(str) -> vals dict
        frmnos_str, frmnos_int = set(), set()
        empty_frm_rows = 0
        invalid_sex_rows = 0
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {}
            for h, f in colmap.items():
                val = row[idx[h]] if idx[h] < len(row) else None
                rec[f] = self._map_sex(val) if h == "SEX" else self._to_text(val)

            # سطر فاضي بالكامل
            if not any(rec.values()):
                continue

            # FRMNO إجباري
            fno = rec.get("frmno")
            if not fno:
                empty_frm_rows += 1
                continue

            # اعتبره "invalid" فقط إذا كانت الخلية غير فاضية فعلاً لكن الماب رجّع False
            sex_cell = None
            if "SEX" in idx and idx["SEX"] < len(row):
                sex_cell = row[idx["SEX"]]
            if sex_cell not in (None, "") and str(sex_cell).strip() != "":
                if rec.get("sex") is False:
                    invalid_sex_rows += 1

            # normalize frmno
            fno = str(fno).strip()
            rec["frmno"] = fno

            # de-duplicate: keep FIRST occurrence
            if fno not in unique_by_frmno:
                unique_by_frmno[fno] = rec
                frmnos_str.add(fno)
                if fno.isdigit():
                    try:
                        frmnos_int.add(int(fno))
                    except Exception:
                        pass

            count += 1
            if count % LOG_EVERY_ROWS == 0:
                _logger.info("ACMST import: pre-read progress %s rows (%.2fs)", count, time.time() - t0)

        _logger.info(
            "ACMST import: pre-read done rows=%s unique_frmno=%s (empty_frm=%s invalid_sex=%s) (%.2fs)",
            count, len(unique_by_frmno), empty_frm_rows, invalid_sex_rows, time.time() - t0
        )

        Student = self.env["acmst.student"].sudo().with_context(
            tracking_disable=True, mail_create_nolog=True, mail_notrack=True
        )

        # --- Robust match of existing (string OR integer) ---
        existing_map = {}  # frmno(str) -> id
        conds = []
        if frmnos_str:
            conds.append(("frmno", "in", list(frmnos_str)))
        if frmnos_int:
            conds.append(("frmno", "in", list(frmnos_int)))
        if conds:
            domain = conds[0]
            for extra in conds[1:]:
                domain = ["|", domain, extra]
            for stu in Student.search(domain).read(["id", "frmno"]):
                key = self._to_text(stu["frmno"])  # normalize to string
                if key:
                    existing_map[key] = stu["id"]
        _logger.info("ACMST import: existing matched=%s (%.2fs)", len(existing_map), time.time() - t0)

        # --- Split into creates / updates ---
        to_create, updates = [], []
        for fno, rec in unique_by_frmno.items():
            ex_id = existing_map.get(fno)
            if ex_id:
                updates.append((ex_id, rec))
            else:
                to_create.append(rec)

        _logger.info(
            "ACMST import: will create=%s update=%s (deduped_in_file=%s) (%.2fs)",
            len(to_create), len(updates), count - len(unique_by_frmno), time.time() - t0
        )

        created = updated = processed = 0
        errors = []

        # --- Batch creates ---
        while to_create:
            chunk = to_create[:BATCH_CREATE]
            del to_create[:BATCH_CREATE]
            try:
                with self.env.cr.savepoint():
                    Student.create(chunk)
                created += len(chunk)
            except Exception:
                # Per-row fallback
                for vals in chunk:
                    try:
                        with self.env.cr.savepoint():
                            Student.create(vals)
                        created += 1
                    except Exception as ee:
                        errors.append(_("Create failed (FRMNO %s): %s") % (vals.get("frmno"), ee))
            processed += len(chunk)
            if processed >= COMMIT_EVERY:
                self.env.cr.commit()
                _logger.info("ACMST import: progress created=%s updated=%s (%.2fs)",
                             created, updated, time.time() - t0)
                processed = 0

        # --- Updates ---
        for rec_id, vals in updates:
            try:
                with self.env.cr.savepoint():
                    Student.browse(rec_id).write(vals)
                updated += 1
            except Exception as e:
                errors.append(_("Update failed (ID %s / FRMNO %s): %s") %
                              (rec_id, vals.get("frmno"), e))
            processed += 1
            if processed >= COMMIT_EVERY:
                self.env.cr.commit()
                _logger.info("ACMST import: progress created=%s updated=%s (%.2fs)",
                             created, updated, time.time() - t0)
                processed = 0

        self.env.cr.commit()
        _logger.info("ACMST import: DONE created=%s updated=%s errors=%s (%.2fs)",
                     created, updated, len(errors), time.time() - t0)

        # --- UI notification ---
        extra = []
        if empty_frm_rows:
            extra.append(_("empty FRMNO rows skipped: %s") % empty_frm_rows)
        if invalid_sex_rows:
            extra.append(_("rows with invalid SEX: %s") % invalid_sex_rows)
        msg = _("Import finished: %(c)s created, %(u)s updated", c=created, u=updated)
        if extra:
            msg += "\n" + " • ".join(extra)
        if errors:
            msg += _("\nSome rows were skipped:\n- ") + "\n- ".join(errors[:10])
            if len(errors) > 10:
                msg += _("\n(and %s more...)") % (len(errors) - 10)

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Students import"),
                "message": msg,
                "type": "success" if not errors else "warning",
                "sticky": False,
                "next": {"type": "ir.actions.act_window_close"},
            },
        }
